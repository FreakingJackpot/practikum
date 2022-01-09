from io import BytesIO

from openpyxl import load_workbook
from catalog.models import Category, Product, Color, Vendor, Attribute, AttributeValue


class ExcelProductImporter:
    PRODUCT_FIELDS = ('Артикул', 'Название', 'Производитель', 'Описание', 'Цена', 'Цена со скидкой', 'Цвет', 'Активен')

    def __init__(self, file):
        self.file = file
        self.name__color_ids_map = {color.name: color.id for color in Color.objects.all()}

    def run(self):
        wb = load_workbook(filename=BytesIO(self.file.read()), read_only=True)

        for category_name in wb.sheetnames:
            category = self.__get_model_obj(Category, name=category_name)
            sheet = wb[category.name]

            error = self.__process_sheet(sheet, category)
            if error:
                return error

    @staticmethod
    def __get_model_obj(model, **kwargs):
        obj, _ = model.objects.get_or_create(**kwargs)
        return obj

    def __process_sheet(self, sheet, category):
        products_to_update = []
        values_to_create = []
        values_to_update = []
        values_ids_to_delete = []
        labels = []

        rows = iter(sheet.rows)
        row = next(rows)

        for cell in row:
            labels.append(cell.value.strip())

        attributes = self._get_attributes(labels, category)

        for row in rows:
            item = {}.fromkeys(labels, None)

            for cell in enumerate(row):
                if cell[0] < len(labels):
                    item[labels[cell[0]]] = cell[1].value

            # Проверка на удаленное значение
            if item == {}.fromkeys(labels, None):
                break

            manufacturer = self.__get_model_obj(Vendor, name=item['Производитель'])
            colors_ids = self._get_colors(item)

            try:
                product = self._create_or_update_product(item, category, manufacturer, colors_ids, products_to_update)
            except ValueError as e:
                return f'Не корректная цена или скидочная цена у продукта категории {category.name} ' \
                       f'c артикулом {item["Артикул"]}'

            self._create_or_update_attr_value(product, attributes, item, values_to_create, values_to_update,
                                              values_ids_to_delete)

        Product.objects.bulk_update(products_to_update,
                                    ('price', 'vendor', 'description', 'discount_price', 'active'))

        AttributeValue.objects.bulk_create(values_to_create)
        AttributeValue.objects.bulk_update(values_to_update, ('value',))
        AttributeValue.objects.filter(id__in=values_ids_to_delete).delete()

    def _get_attributes(self, labels, category):
        attributes_labels = set(filter(lambda x: x not in self.PRODUCT_FIELDS, labels))
        attributes_map = {attr.name: attr for attr in Attribute.objects.filter(category=category)}

        to_delete = set(attributes_map.keys()) - set(attributes_labels)
        for label in to_delete:
            AttributeValue.objects.filter(attribute__category=category, attribute=attributes_map[label]).delete()
            category.attribute.remove(attributes_map[label])
            attributes_map.pop(label)

        to_create = set(attributes_labels) - set(attributes_map.keys())
        new_attrs = []
        for label in to_create:
            attr, _ = Attribute.objects.get_or_create(name=label)
            attributes_map[label] = attr
            new_attrs.append(attr)
        category.attribute.add(*new_attrs)

        return attributes_map.values()

    def _create_or_update_product(self, item, category, manufacturer, colors_ids, products_to_update):
        if not (item['Цена со скидкой'] or item['Цена']):
            raise ValueError

        product = Product.objects.prefetch_related('color').filter(vendor_code=item['Артикул']).first()
        new_colors_ids = set(colors_ids)

        if product:
            updates = self._check_product_attr_change(product, item, manufacturer)

            current_colors_ids = set(color.id for color in product.color.all())

            colors_to_remove = current_colors_ids - new_colors_ids
            if colors_to_remove:
                product.color.remove(*colors_to_remove)

            colors_to_add = new_colors_ids - current_colors_ids
            if colors_to_add:
                product.color.add(*colors_to_add)

            if updates:
                products_to_update.append(product)
        else:
            defaults = {'vendor_code': item['Артикул'], 'name': item['Название'], 'category': category,
                        'description': item['Описание'],
                        'price': item['Цена'], 'discount_price': item['Цена со скидкой'],
                        'active': True if item['Активен'] == 'Да' else False, 'vendor': manufacturer}

            product = Product.objects.create(**defaults)
            product.color.add(*colors_ids)

        return product

    @staticmethod
    def _check_product_attr_change(product, item, manufacturer):
        updates = False
        if item['Описание'] != product.description:
            product.description = item['Описание']
            updates = True
        if item['Цена'] != product.price:
            product.price = item['Цена']
            updates = True
        if item['Цена со скидкой'] != product.discount_price:
            product.discount_price = item['Цена со скидкой']
            updates = True
        activity = True if item['Активен'].capitalize() == 'Да' else False
        if activity != product.active:
            product.active = activity
            updates = True
        if manufacturer != product.vendor:
            product.vendor = manufacturer
            updates = True
        return updates

    @staticmethod
    def _create_or_update_attr_value(product, attributes, item, values_to_create, values_to_update,
                                     values_to_delete):

        current_attrs_values = AttributeValue.objects.select_related('attribute').filter(product=product)
        attribute__value_map = {value.attribute.name: value for value in current_attrs_values}

        for attribute in attributes:
            value = item.get(attribute.name, None)
            attribute_value = attribute__value_map.get(attribute.name, None)

            defaults = {'product': product, 'attribute': attribute, 'value': value}

            if attribute_value:
                if value:
                    attribute_value.value = value
                    values_to_update.append(attribute_value)
                else:
                    values_to_delete.append(attribute_value.id)
            elif value:
                values_to_create.append(AttributeValue(**defaults))

    def _get_colors(self, item):
        colors_ids = []
        for name in item['Цвет'].split('/'):
            cleaned_name = name.strip()
            if cleaned_name in self.name__color_ids_map:
                colors_ids.append(self.name__color_ids_map[cleaned_name])
            else:
                color = Color.objects.create(name=cleaned_name)
                colors_ids.append(color.id)
        return colors_ids
